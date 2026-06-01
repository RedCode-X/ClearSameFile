"""Main application window — coordinates panels, backup system, and new UX flow."""

import os
from typing import List

from PySide6.QtWidgets import (
    QMainWindow, QSplitter, QMenu, QStatusBar,
    QMessageBox, QFileDialog, QWidget, QVBoxLayout, QDialog,
    QLabel, QTextEdit, QDialogButtonBox, QTableWidget, QTableWidgetItem,
    QHeaderView, QAbstractItemView, QHBoxLayout, QPushButton,
)
from PySide6.QtCore import Qt, QThread
from PySide6.QtGui import QAction, QColor, QBrush

from core.scanner import ScanConfig
from core.worker import ScanWorker
from core.deduper import DuplicateGroup, select_best_file, find_trusted_indices
from core import backup
from ui.scan_panel import ScanPanel
from ui.result_panel import ResultPanel
from ui.detail_panel import DetailPanel
from ui.dialogs import ConfirmDeleteDialog, EmptyDirCleanDialog, format_size
from ui.restore_panel import RestoreDialog
from ui.history_dialog import HistoryDialog
from utils.config import (
    save_recent_directories, save_history_entry,
    load_window_geometry, save_window_geometry,
    load_trusted_directories, save_trusted_directories,
)
from utils.exporter import export_csv


DELETE_BG = QColor(255, 235, 235)
DELETE_FG = QColor(180, 40, 40)
KEEP_BG = QColor(230, 250, 230)
KEEP_FG = QColor(30, 130, 30)


class DeletePreviewDialog(QDialog):
    """Table-based preview of files to be deleted / kept, with export support."""

    def __init__(self, groups: List[DuplicateGroup], checked_files: dict, parent=None):
        super().__init__(parent)
        self.setWindowTitle("删除预览")
        self.setMinimumSize(900, 550)

        layout = QVBoxLayout(self)

        total_files = sum(len(v) for v in checked_files.values())
        total_size = 0
        for group_idx, paths in checked_files.items():
            if group_idx < len(groups):
                for f in groups[group_idx].files:
                    if f.path in paths:
                        total_size += f.size

        layout.addWidget(QLabel(
            f"已标记 <b style='color:#c9302c'>{total_files}</b> 个文件待删除, "
            f"共 <b>{format_size(total_size)}</b> — 文件将移至备份目录，可随时恢复。"
        ))

        # Table
        self.table = QTableWidget()
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(["组号", "状态", "文件名", "路径", "大小"])
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setAlternatingRowColors(False)
        self.table.verticalHeader().setVisible(False)
        self.table.setSortingEnabled(True)

        hh = self.table.horizontalHeader()
        hh.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        self.table.setColumnWidth(0, 50)
        hh.setSectionResizeMode(1, QHeaderView.ResizeMode.Fixed)
        self.table.setColumnWidth(1, 55)
        hh.setSectionResizeMode(2, QHeaderView.ResizeMode.Interactive)
        self.table.setColumnWidth(2, 200)
        hh.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        hh.setSectionResizeMode(4, QHeaderView.ResizeMode.Fixed)
        self.table.setColumnWidth(4, 85)

        # Populate rows
        row = 0
        for group_idx, paths in sorted(checked_files.items()):
            if group_idx >= len(groups):
                continue
            g = groups[group_idx]
            for f in g.files:
                is_delete = f.path in paths
                self.table.insertRow(row)

                # Group #
                item = QTableWidgetItem(str(group_idx + 1))
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.table.setItem(row, 0, item)

                # Status
                status = "删除" if is_delete else "保留"
                item = QTableWidgetItem(status)
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.table.setItem(row, 1, item)

                # Filename
                self.table.setItem(row, 2, QTableWidgetItem(f.name))

                # Path
                self.table.setItem(row, 3, QTableWidgetItem(f.dir))

                # Size
                item = QTableWidgetItem(format_size(f.size))
                item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                self.table.setItem(row, 4, item)

                # Color row
                bg = DELETE_BG if is_delete else KEEP_BG
                fg = DELETE_FG if is_delete else KEEP_FG
                for col in range(5):
                    self.table.item(row, col).setBackground(QBrush(bg))
                    self.table.item(row, col).setForeground(QBrush(fg))

                self.table.setRowHeight(row, 24)
                row += 1

        layout.addWidget(self.table)

        # Bottom: summary + buttons
        bottom = QHBoxLayout()
        bottom.addWidget(QLabel(f"共 {row} 个文件"))

        self.export_csv_btn = QPushButton("导出 CSV...")
        self.export_csv_btn.clicked.connect(lambda: self._export("csv"))
        self.export_xlsx_btn = QPushButton("导出 Excel...")
        self.export_xlsx_btn.clicked.connect(lambda: self._export("xlsx"))

        bottom.addStretch()
        bottom.addWidget(self.export_csv_btn)
        bottom.addWidget(self.export_xlsx_btn)
        close_btn = QPushButton("关闭")
        close_btn.clicked.connect(self.close)
        bottom.addWidget(close_btn)
        layout.addLayout(bottom)

    def _export(self, fmt: str):
        if fmt == "csv":
            default_name = "delete_preview.csv"
            filter_str = "CSV 文件 (*.csv)"
        else:
            default_name = "delete_preview.xlsx"
            filter_str = "Excel 文件 (*.xlsx)"

        path, _ = QFileDialog.getSaveFileName(self, "导出删除预览", default_name, filter_str)
        if not path:
            return

        if fmt == "csv":
            self._write_csv(path)
        else:
            self._write_xlsx(path)

    def _write_csv(self, filepath: str):
        import csv
        with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(["组号", "状态", "文件名", "路径", "大小"])
            for row in range(self.table.rowCount()):
                w.writerow([
                    self.table.item(row, 0).text(),
                    self.table.item(row, 1).text(),
                    self.table.item(row, 2).text(),
                    self.table.item(row, 3).text(),
                    self.table.item(row, 4).text(),
                ])
        QMessageBox.information(self, "导出完成", f"已导出到:\n{filepath}")

    def _write_xlsx(self, filepath: str):
        try:
            import openpyxl
        except ImportError:
            QMessageBox.warning(
                self, "缺少依赖",
                "导出 Excel 需要 openpyxl 库。\n请运行: pip install openpyxl\n\n将改为导出 CSV 格式。"
            )
            csv_path = filepath.rsplit(".", 1)[0] + ".csv"
            self._write_csv(csv_path)
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "删除预览"
        ws.append(["组号", "状态", "文件名", "路径", "大小"])

        delete_fill = openpyxl.styles.PatternFill(start_color="FFEBEB", end_color="FFEBEB", fill_type="solid")
        keep_fill = openpyxl.styles.PatternFill(start_color="E6FAE6", end_color="E6FAE6", fill_type="solid")

        for row in range(self.table.rowCount()):
            row_data = [self.table.item(row, col).text() for col in range(5)]
            ws.append(row_data)
            fill = delete_fill if row_data[1] == "删除" else keep_fill
            for col in range(1, 6):
                ws.cell(row=row + 2, column=col).fill = fill

        wb.save(filepath)
        QMessageBox.information(self, "导出完成", f"已导出到:\n{filepath}")


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self._results: List[DuplicateGroup] = []
        self._checked_files: dict[int, set[str]] = {}  # group_index → {paths to delete}
        self._current_group_idx: int = -1
        self._thread: QThread | None = None
        self._worker: ScanWorker | None = None

        self.setWindowTitle("ClearSameFile — 重复文件清理工具")
        self.setMinimumSize(950, 580)

        self._build_menu()
        self._build_ui()
        self._build_statusbar()
        self._restore_geometry()

        # Restore trusted directories from config
        trusted = load_trusted_directories()
        if trusted:
            self.scan_panel.set_trusted_directories(trusted)

    # ==================== UI Construction ====================

    def _build_menu(self):
        mb = self.menuBar()

        file_menu = mb.addMenu("文件(&F)")
        file_menu.addAction(QAction("导出报告为 CSV...", self, triggered=self._export_report))
        file_menu.addSeparator()
        file_menu.addAction(QAction("退出(&X)", self, triggered=self.close))

        tools_menu = mb.addMenu("工具(&T)")
        tools_menu.addAction(QAction("扫描历史记录...", self, triggered=self._open_history))
        tools_menu.addAction(QAction("文件恢复 — 从备份还原...", self, triggered=self._open_restore))
        tools_menu.addSeparator()
        tools_menu.addAction(QAction("切换深色/浅色主题", self, triggered=self._toggle_theme))

        help_menu = mb.addMenu("帮助(&H)")
        help_menu.addAction(QAction("关于", self, triggered=self._show_about))

    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        root = QVBoxLayout(central)
        root.setContentsMargins(4, 4, 4, 4)

        h_split = QSplitter(Qt.Orientation.Horizontal)

        self.scan_panel = ScanPanel()
        self.scan_panel.setMinimumWidth(260)
        self.scan_panel.setMaximumWidth(380)

        v_split = QSplitter(Qt.Orientation.Vertical)

        self.result_panel = ResultPanel()
        self.detail_panel = DetailPanel()

        v_split.addWidget(self.result_panel)
        v_split.addWidget(self.detail_panel)
        v_split.setStretchFactor(0, 3)
        v_split.setStretchFactor(1, 2)

        h_split.addWidget(self.scan_panel)
        h_split.addWidget(v_split)
        h_split.setStretchFactor(0, 1)
        h_split.setStretchFactor(1, 3)

        root.addWidget(h_split)

        # Signals
        self.scan_panel.scan_requested.connect(self._start_scan)
        self.scan_panel.stop_requested.connect(self._stop_scan)
        self.scan_panel.trusted_dirs_changed.connect(self._on_trusted_dirs_changed)

        self.result_panel.group_selected.connect(self._on_group_selected)
        self.result_panel.strategy_changed.connect(self._on_strategy_changed)
        self.result_panel.delete_requested.connect(self._delete_selected)
        self.result_panel.preview_requested.connect(self._preview_delete)
        self.result_panel.export_requested.connect(self._export_report)
        self.result_panel.clean_empty_requested.connect(self._clean_empty_dirs)

        self.detail_panel.deletion_changed.connect(self._on_detail_changed)

    def _build_statusbar(self):
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.status_bar.showMessage("就绪 — 添加目录后点击「开始扫描」")

    def _restore_geometry(self):
        geo = load_window_geometry()
        if geo:
            if "x" in geo:
                self.move(geo.get("x", 100), geo.get("y", 100))
            self.resize(geo.get("w", 1100), geo.get("h", 700))
        else:
            self.resize(1100, 700)

    def _save_geometry(self):
        save_window_geometry({"x": self.x(), "y": self.y(), "w": self.width(), "h": self.height()})

    def closeEvent(self, event):
        self._save_geometry()
        if self._thread and self._thread.isRunning():
            self._worker.cancel()
            self._thread.quit()
            self._thread.wait(3000)
        super().closeEvent(event)

    # ==================== Scan Lifecycle ====================

    def _start_scan(self, config: ScanConfig):
        self._results = []
        self._checked_files.clear()
        self._current_group_idx = -1
        self.result_panel.set_groups([])
        self.detail_panel.set_group(None)
        self.status_bar.showMessage("正在扫描...")

        save_recent_directories(config.directories)
        save_trusted_directories(self.scan_panel.get_trusted_directories())

        self._thread = QThread()
        self._worker = ScanWorker(config)
        self._worker.moveToThread(self._thread)

        self._thread.started.connect(self._worker.run)
        self._worker.progress.connect(self.scan_panel.on_scan_progress)
        self._worker.files_found.connect(self._on_files_found)
        self._worker.finished.connect(self._on_scan_finished)
        self._worker.cancelled.connect(self._on_scan_cancelled)
        self._worker.error.connect(self._on_scan_error)
        self._worker.finished.connect(self._thread.quit)
        self._worker.cancelled.connect(self._thread.quit)
        self._worker.error.connect(self._thread.quit)
        self._thread.finished.connect(self._thread.deleteLater)

        self._thread.start()

    def _stop_scan(self):
        if self._worker:
            self._worker.cancel()

    def _on_files_found(self, count: int):
        self.status_bar.showMessage(f"扫描到 {count:,} 个文件，正在比对重复...")

    def _on_scan_finished(self, groups: list):
        self._results = groups
        self.result_panel.set_groups(groups)
        self.detail_panel.set_group(None)
        self._current_group_idx = -1

        if not groups:
            self.scan_panel.on_scan_finished(0, 0, 0)
            self.status_bar.showMessage("扫描完成 — 未发现重复文件")
            QMessageBox.information(self, "扫描结果", "未发现重复文件。")
            return

        # Auto-apply default strategy (keep_newest) to all groups
        self._apply_strategy_to_all(self.result_panel.get_current_strategy())

        total_files = sum(len(g.files) for g in groups)
        duplicate_files = total_files - len(groups)
        wasted_size = sum(g.wasted_size for g in groups)

        self.scan_panel.on_scan_finished(total_files, duplicate_files, wasted_size)
        total_checked = sum(len(v) for v in self._checked_files.values())
        self.status_bar.showMessage(
            f"扫描完成 — {len(groups)} 组重复, {total_checked} 个文件已标记删除, "
            f"可释放 {format_size(wasted_size)}"
        )

        try:
            import datetime
            dirs = [self.scan_panel.dir_list.item(i).text()
                    for i in range(self.scan_panel.dir_list.count())]
            save_history_entry({
                "time": datetime.datetime.now().isoformat(),
                "groups": len(groups),
                "duplicate_files": duplicate_files,
                "wasted_size": wasted_size,
                "directories": dirs,
            })
        except Exception:
            pass

    def _on_scan_cancelled(self):
        self.scan_panel.on_scan_finished(
            sum(len(g.files) for g in self._results) if self._results else 0,
            0, 0,
        )
        self.status_bar.showMessage("扫描已停止")

    def _on_scan_error(self, msg: str):
        self.scan_panel.on_scan_error(msg)
        self.status_bar.showMessage("扫描出错")

    # ==================== Strategy ====================

    def _on_trusted_dirs_changed(self):
        save_trusted_directories(self.scan_panel.get_trusted_directories())
        if not self._results:
            return
        self._apply_strategy_to_all(self.result_panel.get_current_strategy())
        if self._current_group_idx >= 0 and self._current_group_idx < len(self._results):
            checked = self._checked_files.get(self._current_group_idx, set())
            self.detail_panel.set_group(
                self._results[self._current_group_idx],
                checked_files=checked,
            )

    def _on_strategy_changed(self, strategy: str):
        if strategy == "reset":
            self._reset_all_selections()
            return
        self._apply_strategy_to_all(strategy)
        if self._current_group_idx >= 0:
            checked = self._checked_files.get(self._current_group_idx, set())
            self.detail_panel.set_group(
                self._results[self._current_group_idx],
                checked_files=checked,
            )

    def _apply_strategy_to_all(self, strategy: str):
        """Apply the given strategy to every group, rebuilding _checked_files."""
        self._checked_files.clear()
        for i, g in enumerate(self._results):
            if len(g.files) < 2:
                continue
            delete_paths = self._compute_strategy_files(g, strategy)
            if delete_paths:
                self._checked_files[i] = delete_paths
        self._refresh_result_counts()

    def _get_trusted_dirs(self) -> list:
        return self.scan_panel.get_trusted_directories()

    def _compute_strategy_files(self, group: DuplicateGroup, strategy: str) -> set[str]:
        """Return set of file paths to delete for one group based on strategy.

        Trusted directories take priority: ALL files inside trusted dirs are
        kept unconditionally. Only files outside trusted dirs are marked for
        deletion. If none are in trusted dirs, the normal strategy applies
        (keep 1 best file).
        """
        sorted_files = sorted(group.files, key=lambda f: f.path)
        if len(sorted_files) < 2:
            return set()

        trusted = self._get_trusted_dirs()
        if trusted:
            trusted_idx = set(find_trusted_indices(sorted_files, trusted))
            if trusted_idx:
                # Keep ALL trusted files, delete all non-trusted files
                return {f.path for i, f in enumerate(sorted_files) if i not in trusted_idx}

        keep_idx = select_best_file(sorted_files, strategy)
        if keep_idx < 0:
            return set()

        return {f.path for i, f in enumerate(sorted_files) if i != keep_idx}

    def _reset_all_selections(self):
        self._checked_files.clear()
        self._refresh_result_counts()
        if self._current_group_idx >= 0:
            self.detail_panel.set_group(
                self._results[self._current_group_idx],
                checked_files=set(),
            )
        self.status_bar.showMessage("已重置所有选择")

    def _refresh_result_counts(self):
        """Update the '待删' column in the result panel."""
        row_counts = {i: len(paths) for i, paths in self._checked_files.items()}
        self.result_panel.update_delete_counts(row_counts)

    # ==================== Group Selection ====================

    def _on_group_selected(self, group_idx: int):
        if 0 <= group_idx < len(self._results):
            self._current_group_idx = group_idx
            checked = self._checked_files.get(group_idx, set())
            self.detail_panel.set_group(self._results[group_idx], checked_files=checked)

    def _on_detail_changed(self):
        """User manually toggled a checkbox in the detail panel."""
        if self._current_group_idx < 0:
            return
        files = self.detail_panel.get_selected_files()
        if files:
            self._checked_files[self._current_group_idx] = set(files)
        elif self._current_group_idx in self._checked_files:
            del self._checked_files[self._current_group_idx]
        self._refresh_result_counts()

    # ==================== Delete ====================

    def _preview_delete(self):
        if not self._checked_files:
            return
        dlg = DeletePreviewDialog(self._results, self._checked_files, self)
        dlg.exec()

    def _delete_selected(self):
        if not self._checked_files:
            return

        # Collect files to delete
        files_to_delete: List[str] = []
        total_size = 0
        for group_idx, paths in self._checked_files.items():
            if group_idx < len(self._results):
                for f in self._results[group_idx].files:
                    if f.path in paths:
                        files_to_delete.append(f.path)
                        total_size += f.size

        if not files_to_delete:
            return

        # Confirm
        dlg = ConfirmDeleteDialog(len(files_to_delete), total_size, self)
        dlg.exec()
        if not dlg.confirmed:
            return

        # Move to backup
        self.status_bar.showMessage(f"正在备份并删除 {len(files_to_delete)} 个文件...")
        result = backup.backup_files(files_to_delete)

        success = result["total_files"]
        failed = len(result.get("failed", []))

        self._refresh_results(set(files_to_delete))

        self.status_bar.showMessage(
            f"删除完成 — {success} 个文件已移至备份目录"
            + (f", 失败 {failed} 个" if failed else "")
        )

        if failed > 0:
            QMessageBox.warning(
                self, "删除结果",
                f"成功备份 {success} 个文件到:\n{result['session_dir']}\n\n"
                f"{failed} 个文件删除失败（可能没有权限）。\n\n"
                f"恢复方法: 菜单 → 工具 → 文件恢复"
            )
        else:
            QMessageBox.information(
                self, "删除完成",
                f"已将 {success} 个文件移至备份目录:\n{result['session_dir']}\n\n"
                f"如需恢复: 菜单 → 工具 → 文件恢复"
            )

    def _refresh_results(self, deleted_paths: set[str]):
        """Remove deleted files from results and UI."""
        new_results = []
        for g in self._results:
            remaining = [f for f in g.files if f.path not in deleted_paths]
            if len(remaining) > 1:
                g.files = remaining
                g.total_size = sum(f.size for f in remaining)
                g.wasted_size = g.total_size - remaining[0].size
                new_results.append(g)

        self._results = new_results
        self._checked_files.clear()
        self._current_group_idx = -1
        self.result_panel.set_groups(new_results)
        self.detail_panel.set_group(None)

        total_files = sum(len(g.files) for g in new_results)
        duplicate_files = total_files - len(new_results) if new_results else 0
        wasted_size = sum(g.wasted_size for g in new_results)
        self.scan_panel.on_scan_finished(total_files, duplicate_files, wasted_size)

        if new_results:
            self._apply_strategy_to_all(self.result_panel.get_current_strategy())

    # ==================== Clean Empty Dirs ====================

    def _clean_empty_dirs(self):
        scan_dirs = [
            self.scan_panel.dir_list.item(i).text()
            for i in range(self.scan_panel.dir_list.count())
        ]
        if not scan_dirs:
            QMessageBox.information(self, "提示", "请先添加扫描目录。")
            return

        self.status_bar.showMessage("正在扫描空文件夹...")
        empty_dirs = backup.find_empty_dirs(scan_dirs)
        if not empty_dirs:
            self.status_bar.showMessage("未发现空文件夹")
            QMessageBox.information(self, "清理空目录", "扫描范围内未发现空文件夹。")
            return

        # Show preview
        dlg = EmptyDirCleanDialog(empty_dirs, self)
        dlg.exec()
        if not dlg.confirmed:
            self.status_bar.showMessage("已取消")
            return

        removed = backup.remove_empty_dirs(scan_dirs)
        self.status_bar.showMessage(
            f"已移除 {len(removed)} 个空文件夹"
        )
        QMessageBox.information(
            self, "清理完成",
            f"已移除 {len(removed)} 个空文件夹。"
        )

    # ==================== History ====================

    def _open_history(self):
        dlg = HistoryDialog(self)
        dlg.reload_requested.connect(self._load_directories_from_history)
        dlg.exec()

    def _load_directories_from_history(self, dirs: list):
        self.scan_panel.dir_list.clear()
        for d in dirs:
            if d and os.path.isdir(d):
                self.scan_panel.dir_list.addItem(d)
        self.status_bar.showMessage(f"已加载 {len(dirs)} 个历史目录，可开始扫描")

    # ==================== Restore ====================

    def _open_restore(self):
        # Clean empty sessions first
        backup.clean_empty_sessions()
        dlg = RestoreDialog(self)
        dlg.restore_completed.connect(self._on_files_restored)
        dlg.exec()

    def _on_files_restored(self, count: int):
        self.status_bar.showMessage(f"已恢复 {count} 个文件 — 建议重新扫描以更新重复列表")
        QMessageBox.information(
            self, "提示",
            f"已恢复 {count} 个文件。\n\n建议重新扫描目录以更新重复文件列表。"
        )

    # ==================== Export ====================

    def _export_report(self):
        if not self._results:
            QMessageBox.information(self, "提示", "没有重复文件数据可导出。")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "导出报告", "duplicate_files.csv", "CSV 文件 (*.csv)"
        )
        if path:
            try:
                export_csv(self._results, path)
                self.status_bar.showMessage(f"报告已导出到: {path}")
            except Exception as e:
                QMessageBox.critical(self, "导出失败", f"导出时出错: {e}")

    # ==================== Theme ====================

    def _toggle_theme(self):
        from PySide6.QtWidgets import QApplication
        app = QApplication.instance()
        if app.styleSheet():
            app.setStyleSheet("")
        else:
            app.setStyleSheet("""
                QMainWindow, QWidget { background-color: #2b2b2b; color: #ddd; }
                QGroupBox { border: 1px solid #555; border-radius: 4px; margin-top: 8px; padding-top: 12px; font-weight: bold; }
                QGroupBox::title { subcontrol-origin: margin; left: 10px; padding: 0 3px; }
                QTableWidget { background-color: #333; alternate-background-color: #3a3a3a; gridline-color: #555; }
                QHeaderView::section { background-color: #444; padding: 4px; border: 1px solid #555; }
                QPushButton { background-color: #444; border: 1px solid #555; padding: 5px 12px; border-radius: 3px; }
                QPushButton:hover { background-color: #555; }
                QListWidget { background-color: #333; }
                QProgressBar { background-color: #333; border: 1px solid #555; }
                QProgressBar::chunk { background-color: #5cb85c; }
                QStatusBar { background-color: #222; }
                QMenuBar { background-color: #222; }
                QMenuBar::item:selected { background-color: #444; }
                QMenu { background-color: #333; }
                QMenu::item:selected { background-color: #5cb85c; }
                QComboBox { background-color: #444; border: 1px solid #555; padding: 3px; }
                QComboBox QAbstractItemView { background-color: #333; selection-background-color: #5cb85c; }
                QTextEdit { background-color: #333; border: 1px solid #555; }
            """)

    # ==================== About ====================

    def _show_about(self):
        QMessageBox.about(
            self, "关于 ClearSameFile",
            "<h3>ClearSameFile v1.0</h3>"
            "<p>重复文件查找与清理工具</p>"
            "<p>功能: 扫描磁盘中的重复文件，支持按哈希值精确比对，"
            "选择性删除以释放磁盘空间。</p>"
            "<p>安全: 删除的文件自动备份到「我的文档」下的 ClearSameFile_Backup 目录，"
            "可随时通过「工具 → 文件恢复」还原。</p>"
        )
